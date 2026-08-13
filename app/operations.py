from __future__ import annotations

import csv
import io
from datetime import datetime, timedelta, timezone
from typing import Any

from fastapi import HTTPException
from sqlalchemy import select
from sqlalchemy.orm import Session

from .models import BusinessGoal, BusinessRecord, OperatingEntity, Task
from .task_state import record_task_created


def validate_entity(db: Session, entity_type: str, parent_id: int | None, data: dict[str, Any]) -> None:
    parent = db.get(OperatingEntity, parent_id) if parent_id else None
    expected = {"site": "client", "contract": "site", "shift": "site", "complaint": "site"}
    if entity_type in expected and (not parent or parent.entity_type != expected[entity_type]):
        raise HTTPException(422, f"{entity_type} requires parent entity of type {expected[entity_type]}")
    if entity_type == "employee" and data.get("hourly_rate") is not None:
        try:
            if float(data["hourly_rate"]) < 0:
                raise ValueError
        except (TypeError, ValueError):
            raise HTTPException(422, "employee data.hourly_rate must be non-negative")
    if entity_type == "contract" and data.get("monthly_revenue") is None:
        raise HTTPException(422, "contract requires data.monthly_revenue")


def entity_view(row: OperatingEntity) -> dict[str, Any]:
    return {"id": row.id, "entity_type": row.entity_type, "name": row.name, "status": row.status, "parent_id": row.parent_id, "owner": row.owner, "data": row.data, "started_at": row.started_at, "ended_at": row.ended_at}


def business_graph(db: Session) -> dict[str, Any]:
    rows = db.scalars(select(OperatingEntity).order_by(OperatingEntity.id)).all()
    children: dict[int | None, list[OperatingEntity]] = {}
    for row in rows:
        children.setdefault(row.parent_id, []).append(row)

    def node(row: OperatingEntity) -> dict[str, Any]:
        return {**entity_view(row), "children": [node(child) for child in children.get(row.id, [])]}

    return {"roots": [node(row) for row in children.get(None, [])], "unlinked": [entity_view(x) for x in rows if x.parent_id is None and x.entity_type not in {"client", "employee", "vacancy"}]}


def site_economics(db: Session, site_id: int | None = None) -> list[dict[str, Any]]:
    sites_query = select(OperatingEntity).where(OperatingEntity.entity_type == "site")
    if site_id:
        sites_query = sites_query.where(OperatingEntity.id == site_id)
    sites = db.scalars(sites_query).all()
    entities = db.scalars(select(OperatingEntity)).all()
    result = []
    for site in sites:
        linked = [x for x in entities if x.parent_id == site.id]
        revenue = sum(float(x.data.get("monthly_revenue", 0) or 0) for x in linked if x.entity_type == "contract" and x.status == "active")
        payroll = sum(float(x.data.get("payroll_cost", 0) or 0) for x in linked if x.entity_type == "shift")
        materials = float(site.data.get("materials_cost", 0) or 0)
        logistics = float(site.data.get("logistics_cost", 0) or 0)
        penalties = float(site.data.get("penalties", 0) or 0)
        other = float(site.data.get("other_costs", 0) or 0)
        costs = payroll + materials + logistics + penalties + other
        profit = revenue - costs
        margin = round(profit / revenue * 100, 2) if revenue else 0
        complaints = sum(x.entity_type == "complaint" and x.status not in {"resolved", "closed"} for x in linked)
        result.append({"site_id": site.id, "site": site.name, "revenue": revenue, "payroll": payroll, "materials": materials, "logistics": logistics, "penalties": penalties, "other_costs": other, "profit": profit, "margin_percent": margin, "open_complaints": complaints})
    return result


def simulate_site(db: Session, site_id: int | None, revenue_change_percent: float, payroll_change_percent: float, materials_change_percent: float, penalty_change: float) -> dict[str, Any]:
    current_rows = site_economics(db, site_id)
    if site_id and not current_rows:
        raise HTTPException(404, "Site not found")
    current = {
        "revenue": sum(x["revenue"] for x in current_rows),
        "payroll": sum(x["payroll"] for x in current_rows),
        "materials": sum(x["materials"] for x in current_rows),
        "other": sum(x["logistics"] + x["other_costs"] for x in current_rows),
        "penalties": sum(x["penalties"] for x in current_rows),
    }

    def scenario(multiplier: float) -> dict[str, float]:
        revenue = current["revenue"] * (1 + revenue_change_percent * multiplier / 100)
        payroll = current["payroll"] * (1 + payroll_change_percent * multiplier / 100)
        materials = current["materials"] * (1 + materials_change_percent * multiplier / 100)
        penalties = max(0, current["penalties"] + penalty_change * multiplier)
        profit = revenue - payroll - materials - current["other"] - penalties
        return {"revenue": round(revenue, 2), "costs": round(payroll + materials + current["other"] + penalties, 2), "profit": round(profit, 2), "margin_percent": round(profit / revenue * 100, 2) if revenue else 0}

    return {"current": scenario(0), "conservative": scenario(0.7), "base": scenario(1), "optimistic": scenario(1.3), "assumptions": {"revenue_change_percent": revenue_change_percent, "payroll_change_percent": payroll_change_percent, "materials_change_percent": materials_change_percent, "penalty_change": penalty_change}}


def score_tender(data: dict[str, Any]) -> dict[str, Any]:
    factors = {
        "margin": (float(data.get("expected_margin", 0)), 25),
        "fit": (float(data.get("company_fit", 0)), 20),
        "competition": (100 - float(data.get("competition_risk", 50)), 15),
        "contract_risk": (100 - float(data.get("contract_risk", 50)), 15),
        "logistics": (float(data.get("logistics_fit", 0)), 10),
        "staffing": (float(data.get("staffing_fit", 0)), 10),
        "strategic_value": (float(data.get("strategic_value", 0)), 5),
    }
    breakdown = {name: round(max(0, min(100, value)) * weight / 100, 2) for name, (value, weight) in factors.items()}
    score = round(sum(breakdown.values()), 2)
    return {"score": score, "breakdown": breakdown, "recommendation": "prepare" if score >= 80 else "review" if score >= 60 else "skip"}


def goal_progress(row: BusinessGoal) -> dict[str, Any]:
    span = row.target - row.baseline
    progress = 100 if span == 0 and row.current >= row.target else (row.current - row.baseline) / span * 100 if span else 0
    return {"id": row.id, "title": row.title, "status": row.status, "owner": row.owner, "metric": row.metric, "baseline": row.baseline, "target": row.target, "current": row.current, "unit": row.unit, "progress_percent": round(max(0, min(100, progress)), 2), "deadline_at": row.deadline_at, "strategy": row.strategy}


def create_ceo_actions(db: Session) -> list[Task]:
    tasks: list[Task] = []
    economics = site_economics(db)
    for site in economics:
        if site["revenue"] and site["margin_percent"] < 15:
            tasks.append(Task(title=f"Recover margin at {site['site']}", agent_type="finance", priority="high", payload={"site_id": site["site_id"], "reason": "margin_below_15", "metrics": site}))
        if site["open_complaints"] >= 3:
            tasks.append(Task(title=f"Resolve complaint risk at {site['site']}", agent_type="hr", priority="high", payload={"site_id": site["site_id"], "reason": "complaints_threshold", "metrics": site}))
    overdue = db.scalars(select(BusinessRecord).where(BusinessRecord.record_type == "payment", BusinessRecord.status == "overdue")).all()
    if overdue:
        tasks.append(Task(title="Recover overdue customer payments", agent_type="finance", priority="high", payload={"payment_ids": [x.id for x in overdue], "reason": "overdue_payments"}))
    unique: list[Task] = []
    for task in tasks:
        exists = db.scalar(select(Task.id).where(Task.title == task.title, Task.status.in_(["open", "queued", "running", "blocked"])))
        if not exists:
            db.add(task); unique.append(task)
    db.flush()
    for task in unique:
        record_task_created(db, task, actor="ceo", reason="deterministic_ceo_action")
    return unique


CEO_DEVELOPMENT_BACKLOG = (
    {
        "title": "CEO · Развитие сайта: аудит конверсии и контента",
        "agent_type": "marketing",
        "action": "website_growth_review",
        "scope": "website",
    },
    {
        "title": "CEO · Продажи: анализ воронки и следующих действий",
        "agent_type": "sales",
        "action": "sales_pipeline_review",
        "scope": "sales",
    },
    {
        "title": "CEO · Реклама: анализ каналов и маркетинговых гипотез",
        "agent_type": "marketing",
        "action": "marketing_channel_review",
        "scope": "marketing",
    },
    {
        "title": "CEO · Система: анализ качества агентов и процессов",
        "agent_type": "meta_brain",
        "action": "agent_quality_review",
        "scope": "system",
    },
)


def maintain_ceo_development_backlog(
    db: Session,
    *,
    now: datetime | None = None,
    cadence_hours: int = 24,
) -> list[Task]:
    """Keep a safe, finite and recurring CEO development backlog."""
    current_time = now or datetime.now(timezone.utc).replace(tzinfo=None)
    cadence = max(1, min(int(cadence_hours), 7 * 24))
    created: list[Task] = []
    for template in CEO_DEVELOPMENT_BACKLOG:
        latest = db.scalar(
            select(Task).where(Task.title == template["title"]).order_by(Task.id.desc())
        )
        if latest and latest.status in {"open", "queued", "running", "blocked", "failed"}:
            continue
        run_after = current_time
        if latest:
            run_after = max(current_time, latest.run_after + timedelta(hours=cadence))
        task = Task(
            title=template["title"],
            agent_type=template["agent_type"],
            status="queued",
            priority="normal",
            run_after=run_after,
            max_attempts=3,
            payload={
                "action": template["action"],
                "scope": template["scope"],
                "origin": "ceo_continuous_backlog",
                "advisory_only": True,
                "external_actions_require_owner_approval": True,
            },
        )
        db.add(task)
        db.flush()
        record_task_created(db, task, actor="ceo", reason="recurring_development_backlog")
        created.append(task)
    return created


def parse_lead_import(filename: str, content: bytes) -> list[dict[str, Any]]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        text = content.decode("utf-8-sig")
        return [dict(row) for row in csv.DictReader(io.StringIO(text))]
    if lower.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(503, "XLSX import requires openpyxl") from exc
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(x or "").strip() for x in values[0]]
        return [{headers[i]: value for i, value in enumerate(row) if i < len(headers) and headers[i]} for row in values[1:]]
    raise HTTPException(422, "Only CSV and XLSX files are supported")
